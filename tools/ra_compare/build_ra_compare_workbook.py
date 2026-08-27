#!/usr/bin/env python3
"""
Build the RA input comparison workbook (`Compare_RA_<PERIMETER>_...xlsx`).

The Risk team produces this comparison by hand today (`Compare_RA_BCEF_KO_Prod 26Q2.xlsx`).
This script rebuilds the SAME workbook from two `Inputs_RA` files, so the layout the engine
must produce is pinned by something runnable rather than by a screenshot.

It is the LAYOUT REFERENCE for the treatment specified in
`docs/tseadfwd/977/RA_COMPARISON_REPORT_DESIGN.md` — the production job is Scala/POI inside the
`tseadfwd` module; this generator stays the acceptance fixture the job is compared against.

    python3 tools/ra_compare/build_ra_compare_workbook.py \
        --new localRun/tseadfwd/input/Inputs_RA_v2.xlsx \
        --old localRun/tseadfwd/input/Inputs_RA.xlsx \
        --out docs/tseadfwd/977/Compare_RA_reference_model.xlsx

One sheet per (PERIMETER x FWL_TYPE) common to both files, each holding three stacked blocks
(Updated / Previous / Evol) of four segment tables, then one chart per (metric x segment).
"""

import argparse
import re

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- input model

KEY_COLUMNS = ["PERIMETER", "SEGMENT", "RATE_TYPE", "FWL_TYPE", "METRIC"]
SHEET_PATTERN = re.compile(r"^RA[_ ].*", re.IGNORECASE)

# Metric rows of one segment table, in order - the row label is the metric name itself.
#
# The four metrics of the INPUTS_RA file, and nothing else. The manual Risk workbook also
# carries two DERIVED rows - RA = RA STAT + RA FI + RE, and %RA = RA x 12 / -CRD - which the
# business dropped on 2026-08-27 (OPEN_QUESTIONS_977 Q6/Q11/Q12): the report works only from
# the metrics present in the input. That decision is what makes a table 6 rows rather than 8.
#
# The manual file's second label column (Outstanding / Amount PPstat / Amount PPfin / Amount RE,
# next to CRD / RA STAT / RA FI / RE) went with the business review of 2026-08-27: it said the
# same thing twice and cost a column on a 361-column sheet.
BLOCK_ROWS = ["CRD", "RA STAT", "RA FI", "RE"]
METRICS = ["CRD", "RA STAT", "RA FI", "RE"]          # the four metrics of the ticket
CHART_METRICS = ["CRD", "RA STAT", "RA FI", "RE"]

# Metrics the report shows with their sign reversed. CRD is an outstanding and INPUTS_RA carries
# it negative (it is an exposure); the business asked on 2026-08-27 for it to READ positive -
# literally -1 x the input. Presentation only: (-n - -o) / -o == (n - o) / o exactly, so every
# Evol percentage is the number it was before the flip.
NEGATED_METRICS = {"CRD"}


def display_value(metric, value):
    """The value as the report shows it: CRD negated, everything else untouched."""
    if value is None:
        return None
    if metric not in NEGATED_METRICS:
        return value
    # Negating a zero gives -0.0, which Excel renders as "-0": the same number, read as a signed
    # one. Not worth the distraction in a business report.
    return 0.0 if value == 0 else -value


# Segment tables are headed with the segment name AS IT APPEARS IN THE INPUT. The manual Risk
# workbook uses French business wording (Immos / Invest pro / Invest corp / Conso); the business
# chose the input's own names on 2026-08-27 (Q14), so a segment added to INPUTS_RA needs no
# mapping here to appear correctly.
# The display ORDER is configuration (--segment-order); anything not named falls after it,
# alphabetically.
DEFAULT_SEGMENT_ORDER = ["MORTGAGE", "INVEST_PRO", "INVEST_CORP", "CONSO"]

# The manual file names its sheets after the stress leg (+/- 100 bp).
FWL_SHEET_SUFFIX = {"BASELINE": "BASELINE", "STRESS (-)": "-100", "STRESS (+)": "+100"}
FWL_ORDER = ["BASELINE", "STRESS (-)", "STRESS (+)"]

# ------------------------------------------------------------ sheet geometry
# The manual workbook's anchors were 1 / 38 / 75 with a pitch of 8 (header + 6 metric rows +
# blank). Dropping the two derived rows (Q6/Q11) makes a table 6 rows - header + 4 metrics +
# blank - so each block is 4 x 6 = 24 rows of tables instead of 32; dropping the standalone
# month index row (business review 2026-08-27, it repeated the table header's own M1..Mn) takes
# one more row off the top of every block. Derived here rather than hard-coded, so the anchors
# stay correct if the row list changes again.
BLOCK_PITCH = len(BLOCK_ROWS) + 2          # table header + metric rows + one blank
TITLE_TO_FIRST_HEADER = 2                  # title, "En M EUR", then the table header
BLOCK_GAP = 2                              # blank rows between one block and the next title

_BLOCK_HEIGHT = TITLE_TO_FIRST_HEADER + 4 * BLOCK_PITCH + BLOCK_GAP
UPDATED_TITLE_ROW = 1
PREVIOUS_TITLE_ROW = UPDATED_TITLE_ROW + _BLOCK_HEIGHT
EVOL_TITLE_ROW = PREVIOUS_TITLE_ROW + _BLOCK_HEIGHT
UPDATED_FIRST_HEADER = UPDATED_TITLE_ROW + TITLE_TO_FIRST_HEADER
PREVIOUS_FIRST_HEADER = PREVIOUS_TITLE_ROW + TITLE_TO_FIRST_HEADER
EVOL_FIRST_HEADER = EVOL_TITLE_ROW + TITLE_TO_FIRST_HEADER
FIRST_DATA_COL = 2  # column B - column A carries the metric name

YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="808080")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_CRD = "#,##0"
FMT_AMOUNT = "#,##0.00"
FMT_PCT = "0.00%"


def read_ra_file(path):
    """(keys -> monthly series) plus the month column labels, from every RA sheet of a workbook.

    Applies the engine's own two gates: the sheet NAME matches `^RA[_ ].*`, and its CONTENT
    carries the five key columns. Anything else in the workbook is ignored.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    series, months, sheets, skipped = {}, None, [], []
    for name in wb.sheetnames:
        if not SHEET_PATTERN.match(name):
            skipped.append((name, "name does not match ^RA[_ ].*"))
            continue
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            skipped.append((name, "empty sheet"))
            continue
        header = [str(h).strip() if h is not None else "" for h in header]
        if header[:5] != KEY_COLUMNS:
            skipped.append((name, "not an RA table: missing %s" % ", ".join(
                c for c in KEY_COLUMNS if c not in header)))
            continue
        sheet_months = [h for h in header[5:] if h]
        if months is None:
            months = sheet_months
        for row in rows:
            if row is None or row[0] is None:
                continue
            key = tuple(str(v).strip() for v in row[:5])
            series[key] = [(v if isinstance(v, (int, float)) else 0.0)
                           for v in row[5:5 + len(sheet_months)]]
        sheets.append(name)
    wb.close()
    if not series:
        raise SystemExit("no RA sheet could be read from %s" % path)
    return series, months, sheets, skipped


def perimeter_segments(series, perimeter, segment_order):
    """(SEGMENT, RATE_TYPE) pairs of a perimeter, in the configured display order.

    A segment the order does not name still appears - after the named ones, alphabetically -
    so a new segment in INPUTS_RA is never silently dropped from the report."""
    pairs = {(k[1], k[2]) for k in series if k[0] == perimeter}
    known = [(s, r) for s in segment_order for (sg, r) in sorted(pairs) if sg == s]
    rest = sorted(p for p in pairs if p[0] not in segment_order)
    return known + rest


def write_block(ws, first_header_row, title_row, title, series, perimeter, fwl,
                segments, months, n_months):
    """One `Updated` / `Previous` block: the title, the month index row, four segment tables."""
    ws.cell(title_row, 1, title).fill = YELLOW
    ws.cell(title_row, 1).font = Font(bold=True)
    ws.cell(title_row + 1, 1, "En M EUR").font = Font(italic=True)

    for b, (segment, rate_type) in enumerate(segments):
        header = first_header_row + b * BLOCK_PITCH
        h = ws.cell(header, 1, "%s a %s" % (segment, rate_type))
        h.font, h.border = Font(bold=True), BOX
        # Month labels, not calendar dates: an INPUTS_RA sheet carries no as-of date and the
        # business dropped the dates rather than supply one per side (Q3/Q4).
        for i, label in enumerate(months[:n_months]):
            c = ws.cell(header, FIRST_DATA_COL + i, label)
            c.font, c.border, c.alignment = Font(bold=True, size=9), BOX, Alignment(horizontal="center")

        for j, metric in enumerate(BLOCK_ROWS):
            row = header + 1 + j
            ws.cell(row, 1, metric).font = Font(size=9)
            ws.cell(row, 1).border = BOX
            for i in range(n_months):
                cell = ws.cell(row, FIRST_DATA_COL + i)
                values = series.get((perimeter, segment, rate_type, fwl, metric))
                # CRD is written as -1 x the input, so an outstanding reads positive; both blocks
                # are flipped, so the Evol percentages below are unchanged.
                cell.value = display_value(metric, values[i]) if values else None
                cell.number_format = FMT_CRD if metric == "CRD" else FMT_AMOUNT
                cell.font = Font(size=9)


def write_evol_block(ws, segments, n_months, safe_div):
    """The `Evol` block: (Updated - Previous) / Previous, cell by cell, as live formulas."""
    ws.cell(EVOL_TITLE_ROW, 1, "Evol").fill = YELLOW
    ws.cell(EVOL_TITLE_ROW, 1).font = Font(bold=True)
    ws.cell(EVOL_TITLE_ROW + 1, 1, "En M EUR").font = Font(italic=True)

    for b, (segment, rate_type) in enumerate(segments):
        header = EVOL_FIRST_HEADER + b * BLOCK_PITCH
        up_header = UPDATED_FIRST_HEADER + b * BLOCK_PITCH
        prev_header = PREVIOUS_FIRST_HEADER + b * BLOCK_PITCH
        h = ws.cell(header, 1, "%s a %s" % (segment, rate_type))
        h.font, h.border = Font(bold=True), BOX
        for i in range(n_months):
            c = ws.cell(header, FIRST_DATA_COL + i,
                        ws.cell(up_header, FIRST_DATA_COL + i).value)
            c.font, c.border, c.alignment = Font(bold=True, size=9), BOX, Alignment(horizontal="center")

        for j, metric in enumerate(BLOCK_ROWS):
            row = header + 1 + j
            ws.cell(row, 1, metric).font = Font(size=9)
            ws.cell(row, 1).border = BOX
            for i in range(n_months):
                col = FIRST_DATA_COL + i
                letter = get_column_letter(col)
                new_ref = "%s%d" % (letter, up_header + 1 + j)
                old_ref = "%s%d" % (letter, prev_header + 1 + j)
                # The manual file's own formula: =(C5-C42)/C42
                body = "(%s-%s)/%s" % (new_ref, old_ref, old_ref)
                cell = ws.cell(row, col)
                cell.value = "=IFERROR(%s,\"\")" % body if safe_div else "=" + body
                cell.number_format = FMT_PCT
                cell.font = Font(size=9)


def add_charts(ws, segments, n_months, chart_step):
    """One line chart per (metric x segment): the Updated and the Previous curve, superimposed."""
    first_chart_row = EVOL_FIRST_HEADER + len(segments) * BLOCK_PITCH + 3
    max_col = FIRST_DATA_COL + n_months - 1
    for m, metric in enumerate(CHART_METRICS):
        offset = BLOCK_ROWS.index(metric)
        for b, (segment, rate_type) in enumerate(segments):
            chart = LineChart()
            chart.title = "%s - %s" % (metric, segment)
            chart.style = 2
            chart.height, chart.width = 7.5, 12.5
            chart.y_axis.numFmt = FMT_CRD if metric == "CRD" else FMT_AMOUNT
            for block_first, name in ((UPDATED_FIRST_HEADER, "Updated"),
                                      (PREVIOUS_FIRST_HEADER, "Previous")):
                row = block_first + b * BLOCK_PITCH + 1 + offset
                data = Reference(ws, min_col=FIRST_DATA_COL, max_col=max_col,
                                 min_row=row, max_row=row)
                chart.add_data(data, titles_from_data=False, from_rows=True)
                series = chart.series[-1]
                series.tx = SeriesLabel(v=name)
                series.smooth = False
            # month labels are TEXT (M1, M2, …): a strRef keeps Excel from numbering the axis 1..n.
            # They come from this table's own header row - the standalone month index row that used
            # to feed them was dropped as a duplicate (business review 2026-08-27).
            cat_row = UPDATED_FIRST_HEADER + b * BLOCK_PITCH
            categories = AxDataSource(strRef=StrRef("'%s'!$%s$%d:$%s$%d" % (
                ws.title, get_column_letter(FIRST_DATA_COL), cat_row,
                get_column_letter(max_col), cat_row)))
            for series in chart.series:
                series.cat = categories
            chart.x_axis.tickLblSkip = chart_step
            chart.x_axis.tickMarkSkip = chart_step
            anchor = "%s%d" % (get_column_letter(1 + b * 9), first_chart_row + m * 16)
            ws.add_chart(chart, anchor)


def write_info_sheet(wb, args, new_meta, old_meta, common, only_new, only_old, n_months):
    ws = wb.create_sheet("COMPARE INFO", 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    rows = [
        ("RA input comparison", ""),
        ("", ""),
        ("Updated (new) file", args.new),
        ("Previous (old) file", args.old),
        ("Updated RA sheets", ", ".join(new_meta["sheets"])),
        ("Previous RA sheets", ", ".join(old_meta["sheets"])),
        ("Months compared", "%d (%s .. %s), aligned by month INDEX" %
            (n_months, new_meta["months"][0], new_meta["months"][n_months - 1])),
        ("", ""),
        ("Compared keys", "%d perimeter x segment x rate type x FWL type x metric" % len(common)),
        ("Only in Updated (dropped)", ", ".join(sorted({"/".join(k[:3]) for k in only_new})) or "none"),
        ("Only in Previous (dropped)", ", ".join(sorted({"/".join(k[:3]) for k in only_old})) or "none"),
        ("", ""),
        ("%change", "(Updated - Previous) / Previous, per metric, per month, per key"),
        ("Metrics", "CRD, RA STAT, RA FI, RE - as they appear in the input. The manual "
                    "workbook's derived RA and %RA rows are not reproduced (Q6/Q11)."),
        ("Months", "identified as M1..M361; the report carries no calendar dates (Q3/Q4)."),
        ("Values", "used at the scale they appear in INPUTS_RA, with no rescaling (Q16)."),
        ("CRD sign", "shown as -1 x the input value, so an outstanding reads POSITIVE (business "
                     "review 2026-08-27). Both sides are flipped, so every %change is the number "
                     "it would be on the raw values."),
        ("", ""),
        ("Generated by", "tools/ra_compare/build_ra_compare_workbook.py"),
        ("Design", "docs/tseadfwd/977/RA_COMPARISON_REPORT_DESIGN.md"),
    ]
    for i, (a, b) in enumerate(rows, start=1):
        ws.cell(i, 1, a).font = Font(bold=(b == "" and a != ""))
        ws.cell(i, 2, b)
    ws["A1"].font = Font(bold=True, size=14)


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--new", required=True, help="Updated Inputs_RA workbook")
    p.add_argument("--old", required=True, help="Previous Inputs_RA workbook")
    p.add_argument("--out", required=True, help="workbook to write")
    p.add_argument("--segment-order", default=",".join(DEFAULT_SEGMENT_ORDER),
                   help="comma-separated segment display order; anything not named follows, "
                        "alphabetically")
    p.add_argument("--perimeters", default="", help="comma-separated subset; default = all common ones")
    p.add_argument("--chart-step", type=int, default=12, help="x-axis label step of the charts")
    p.add_argument("--raw-div", action="store_true",
                   help="write the manual file's bare =(C5-C42)/C42 (shows #DIV/0!) instead of IFERROR")
    args = p.parse_args()

    segment_order = [x.strip().upper() for x in args.segment_order.split(",") if x.strip()]

    new_series, new_months, new_sheets, new_skipped = read_ra_file(args.new)
    old_series, old_months, old_sheets, old_skipped = read_ra_file(args.old)

    common = sorted(set(new_series) & set(old_series))
    only_new = sorted(set(new_series) - set(old_series))
    only_old = sorted(set(old_series) - set(new_series))
    if not common:
        raise SystemExit("the two files share no (perimeter, segment, rate type, FWL type, metric) key")
    n_months = min(len(new_months), len(old_months),
                   min(len(v) for v in new_series.values()),
                   min(len(v) for v in old_series.values()))

    perimeters = sorted({k[0] for k in common})
    if args.perimeters:
        wanted = {s.strip().upper() for s in args.perimeters.split(",")}
        perimeters = [p for p in perimeters if p in wanted]

    wb = Workbook()
    wb.remove(wb.active)
    for perimeter in perimeters:
        segments = [sr for sr in perimeter_segments(new_series, perimeter, segment_order)
                    if any(k[:3] == (perimeter, sr[0], sr[1]) for k in common)]
        for fwl in FWL_ORDER:
            if not any(k[0] == perimeter and k[3] == fwl for k in common):
                continue
            ws = wb.create_sheet("%s %s" % (perimeter, FWL_SHEET_SUFFIX.get(fwl, fwl)))
            ws.sheet_view.showGridLines = False
            # No freeze pane: the business asked for it off (2026-08-27). It pinned the Updated
            # block's header over the Previous and Evol blocks once you scrolled down.
            ws.column_dimensions["A"].width = 14
            for i in range(n_months):
                ws.column_dimensions[get_column_letter(FIRST_DATA_COL + i)].width = 11

            write_block(ws, UPDATED_FIRST_HEADER, UPDATED_TITLE_ROW, "Updated", new_series,
                        perimeter, fwl, segments, new_months, n_months)
            write_block(ws, PREVIOUS_FIRST_HEADER, PREVIOUS_TITLE_ROW, "Previous", old_series,
                        perimeter, fwl, segments, old_months, n_months)
            write_evol_block(ws, segments, n_months, safe_div=not args.raw_div)
            add_charts(ws, segments, n_months, args.chart_step)

    write_info_sheet(wb, args,
                     {"sheets": new_sheets, "months": new_months},
                     {"sheets": old_sheets, "months": old_months},
                     common, only_new, only_old, n_months)
    wb.save(args.out)

    print("wrote %s" % args.out)
    print("  sheets      : %s" % ", ".join(s.title for s in wb.worksheets))
    print("  months      : %d" % n_months)
    print("  common keys : %d   only-new %d   only-previous %d"
          % (len(common), len(only_new), len(only_old)))
    for name, reason in new_skipped + old_skipped:
        print("  skipped     : %s (%s)" % (name, reason))


if __name__ == "__main__":
    main()
